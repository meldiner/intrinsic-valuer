from openpyxl import load_workbook
import datetime

def write(db, file):
    wb = load_workbook("template.xlsx")
    worksheet = wb['data']

    end_year = datetime.datetime.now().year 
    start_year = end_year - 10 # we are analysing a history of 10 years

    row = 1
    col = 1
    worksheet.cell(column=col, row=row, value="Year")
    col = col + 1

    for key in sorted(db):
        worksheet.cell(column=col, row=row, value=key)
        col = col + 1

    row = row + 1
    col = 1

    for year in (range(start_year, end_year + 1, 1)):
        year_str = str(year)
        worksheet.cell(column=col, row=row, value=year)
        col = col + 1

        for key in sorted(db):
            if year_str in db[key] and db[key][year_str]:
                worksheet.cell(column=col, row=row, value=db[key][year_str])
            col = col + 1

        row = row + 1
        col = 1

    wb.save(file)